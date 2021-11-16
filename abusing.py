#!/usr/bin/python3

# Se importan los módulos. Se han añadido excepciones para que se instalen directamente con pip los módulos openpyxl, tqdm y requests en caso de no estar instalados.

import sys
import subprocess
import re
import ipaddress
import csv
import json
from time import sleep

try:
    import openpyxl
except:
    print('\nEl módulo openpyxl no está instalado. Vamos a intalarlo.\n')
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

from openpyxl.styles import PatternFill, NamedStyle, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter 

try:
    from tqdm import tqdm
except:
    print('\nEl módulo tqdm no está instalado. Vamos a intalarlo.\n')
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'tqdm'])
    from tqdm import tqdm

try:
    import requests
except:
    print('\nEl módulo requests no está instalado. Vamos a intalarlo.\n')
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'requests'])
    import requests

# logo

logo = """

    ___     __                  _                ___     __                        ____ ____   ____   ____ 
   /   |   / /_   __  __ _____ (_)____   ____ _ /   |   / /_   __  __ _____ ___   /  _// __ \ / __ \ / __ )
  / /| |  / __ \ / / / // ___// // __ \ / __ `// /| |  / __ \ / / / // ___// _ \  / / / /_/ // / / // __  |
 / ___ | / /_/ // /_/ /(__  )/ // / / // /_/ // ___ | / /_/ // /_/ /(__  )/  __/_/ / / ____// /_/ // /_/ / 
/_/  |_|/_.___/ \__,_//____//_//_/ /_/ \__, //_/  |_|/_.___/ \__,_//____/ \___//___//_/    /_____//_____/  
                                      /____/                                                                                                                                                                                                                                                                                                                    
"""

print(logo) 

# API KEY, ruta del archivo, url de abuseIPDB y columnas del archivo.

API_KEY = input('\nIntroduce tu API KEY: ')
file = input(
    '\nIntroduce la ruta del archivo de texto en el que se encuentran las IPs que deseas comprobar (si estás en una máquina Windows doblar las barras invertidas para evitar error. e.g: C:\\\\Users\\\\usuario\\\\archivo.txt): ')
url = 'https://api.abuseipdb.com/api/v2/check'

csv_columns = ['ipAddress', 'isPublic', 'ipVersion', 'isWhitelisted', 'abuseConfidenceScore', 'countryCode',
               'usageType', 'isp', 'domain', 'hostnames', 'totalReports', 'numDistinctUsers', 'lastReportedAt']

# Crear archivo csv

with open("resultados.csv", "a", newline='') as filecsv:
    writer = csv.DictWriter(filecsv, fieldnames=csv_columns)
    writer.writeheader()

# Leer ips de archivo

ip_list = []
with open(file) as f:
    file_item = f.read()
    regex = r'(?:(?:2(?:[0-4][0-9]|5[0-5])|[0-1]?[0-9]?[0-9])\.){3}(?:(?:2([0-4][0-9]|5[0-5])|[0-1]?[0-9]?[0-9]))'
    matches = re.finditer(regex, file_item, re.MULTILINE)

    [ip_list.append(match.group())
     for matchNum, match in enumerate(matches, start=1)]

# Comprobar si hay IPs privadas en la lista y de ser así, sacarlas de ella

for i in ip_list:
    if ipaddress.ip_address(i).is_private is True:
        ip_list.remove(i)
        print('\n' + i + ' is a private IP address, and is only used in internal network environments. Any abusive activity you see coming from an internal IP is either coming from within your network itself, or is the result of an error or misconfiguration. ')

# Consulta en AbuseIPDB y rellenar valores en archivo csv

print('\nSe están comprobando las IPs en AbuseIPDB. Esto puede tardar un poco...')
for i in ip_list:
    headers = {
        'Accept': 'application/json',
        'Key': API_KEY}

    params = {
        'ipAddress': i,
        'maxAgeInDays': '30'}

    r = requests.get(url=url, headers=headers, params=params)
    json_Data = json.loads(r.content)
    json_main = json_Data["data"]
    with open("resultados.csv", "a", newline='') as filecsv:
        writer = csv.DictWriter(filecsv, fieldnames=csv_columns)
        writer.writerow(json_main)
    print('\nComprobando ' + str(i))
    for i in tqdm(range(100)):
        sleep(0.01)
    

# Convertir csv a excel 

csv_data = []
with open('resultados.csv') as file_obj:
    reader = csv.reader(file_obj)
    for row in reader:
        csv_data.append(row)
workbook = openpyxl.Workbook()
sheet = workbook.active
for row in csv_data:
    sheet.append(row)

# Ajustar tamaño de las columnas

MIN_WIDTH = 10
for i, column_cells in enumerate(sheet.columns, start=1):
    width = (
        length
        if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                          for cell in column_cells)) >= MIN_WIDTH
        else MIN_WIDTH
    )
    sheet.column_dimensions[get_column_letter(i)].width = width

# Pintar en rojo IPs reportadas

red_background = PatternFill(bgColor="00FF0000", fill_type ="solid")
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style, stopIfTrue=True)
rule.formula = ['int($K1)>0']
sheet.conditional_formatting.add("A1:O1000", rule)

# Primera fila en negrita 

header = NamedStyle(name="header")
header.font = Font(bold=True)
header_row = sheet[1]
for cell in header_row:
    cell.style = header

# Congelar primera fila

sheet.freeze_panes = "A2"                                

# Añadir filtros

sheet.auto_filter.ref = "A1:M1000"                       
                                                           
workbook.save('resultados.xlsx')



